"""Fatura Bot — Fiş Aktarım Şablon formatında günlük Excel/CSV/XLS çıktı servisi."""

import asyncio
import csv
import io
import re
import shutil
from datetime import datetime, timezone, date
from pathlib import Path
from typing import Optional

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

from src.models.schemas import ReceiptData
from src.utils.logger import logger

from src.services.luca_transformer import (
    MASRAF_HESAP_KODU, KDV_HESAP_KODU, ODEME_HESAP_KODU,
    ODEME_BELGE_TR, fis_to_luca_list,
)


class ExcelService:
    """Fiş Aktarım Şablon formatında Excel dosyası oluşturma servisi."""

    TEMPLATE_COLUMNS = [
        ("Fiş No",         14.15, "@"),
        ("Fiş Tarihi",     13.29, "dd/mm/yyyy"),
        ("Fiş Açıklama",   33.71, "@"),
        ("Hesap Kodu",     18.29, "@"),
        ("Evrak No",       15.15, "@"),
        ("Evrak Tarihi",   13.29, "dd/mm/yyyy"),
        ("Detay Açıklama", 34.71, "@"),
        ("Borç",           11.86, "#,##0.00"),
        ("Alacak",         11.86, "#,##0.00"),
        ("Miktar",         11.86, "#,##0.00000"),
        ("Belge Tr",       16.41, "@"),
        ("Para Birimi",    15.79, "@"),
        ("Kur",            11.86, "#,##0.00000000"),
        ("Döviz Tutarı",   17.55, "#,##0.00"),
    ]

    HEADER_FONT = Font(name="Calibri", bold=True, color="FF000000", size=10, charset=1, family=2)
    HEADER_FILL = PatternFill(patternType="solid", fgColor="FFC0C0C0", bgColor="FFCCCCFF")
    HEADER_ALIGNMENT = Alignment(horizontal="center", vertical="center", wrap_text=True)
    HEADER_ROW_HEIGHT = 21.75
    SHEET_NAME = "Fiş Aktarım Şablon"

    def __init__(self, data_dir: str = "public/daily"):
        self._data_dir = Path(data_dir)
        self._data_dir.mkdir(parents=True, exist_ok=True)
        self._lock = asyncio.Lock()
        self._template_path = Path(__file__).parent.parent / "templates" / "fis_aktarim_sablon.xlsx"

    @property
    def current_filename(self) -> str:
        """Günlük Excel dosyasının sadece adını döndürür (yol bilgisi olmadan)."""
        return self._daily_path().name

    def _daily_path(self, target_date: Optional[date] = None) -> Path:
        d = target_date or date.today()
        return self._data_dir / f"{d.isoformat()}.xlsx"

    @staticmethod
    def _parse_receipt_date(tarih_str: Optional[str]) -> Optional[str]:
        """Tarih string'ini DD/MM/YYYY formatına normalleştir."""
        if not tarih_str:
            return None
        try:
            parts = tarih_str.replace(".", "/").replace("-", "/").split("/")
            if len(parts) == 3:
                d, m, y = int(parts[0]), int(parts[1]), int(parts[2])
                if y < 100:
                    y += 2000
                return f"{d:02d}/{m:02d}/{y:04d}"
        except (ValueError, IndexError):
            pass
        return tarih_str

    def _ensure_workbook(self, file_path: Path) -> Workbook:
        """Mevcut dosyayı aç veya şablondan yenisini oluştur."""
        if file_path.exists():
            try:
                wb = load_workbook(str(file_path))
                if self.SHEET_NAME in wb.sheetnames:
                    return wb
                wb.close()
            except Exception as e:
                logger.warning("Mevcut dosya okunamadı, yeniden oluşturulacak", error=str(e))

        if self._template_path.exists():
            try:
                shutil.copy2(str(self._template_path), str(file_path))
                return load_workbook(str(file_path))
            except Exception as e:
                logger.warning("Şablon kopyalanamadı, sıfırdan oluşturuluyor", error=str(e))

        wb = Workbook()
        ws = wb.active
        ws.title = self.SHEET_NAME
        self._setup_header(ws)
        wb.save(str(file_path))
        return wb

    def _setup_header(self, ws):
        """Şablon başlık satırını birebir oluştur."""
        ws.row_dimensions[1].height = self.HEADER_ROW_HEIGHT

        for col_idx, (title, width, num_fmt) in enumerate(self.TEMPLATE_COLUMNS, 1):
            cell = ws.cell(row=1, column=col_idx, value=title)
            cell.font = self.HEADER_FONT
            cell.fill = self.HEADER_FILL
            cell.alignment = self.HEADER_ALIGNMENT
            cell.number_format = num_fmt
            ws.column_dimensions[get_column_letter(col_idx)].width = width

    def _build_rows(self, data: ReceiptData) -> list[list]:
        """Bir fişten şablon formatında çift-kayıt muhasebe satırları üretir."""
        return fis_to_luca_list(data)

    async def add_row(self, data: ReceiptData, confidence: int, sender: str, source: str = "gemini") -> int:
        """Fiş verisini bugünün dosyasına çift kayıt olarak yaz. İlk satır numarasını döndürür."""
        async with self._lock:
            try:
                file_path = self._daily_path(date.today())
                wb = self._ensure_workbook(file_path)
                ws = wb[self.SHEET_NAME]

                first_row = ws.max_row + 1
                template_rows = self._build_rows(data)

                for row_data in template_rows:
                    row_num = ws.max_row + 1
                    for col_idx, value in enumerate(row_data, 1):
                        cell = ws.cell(row=row_num, column=col_idx, value=value)
                        cell.number_format = self.TEMPLATE_COLUMNS[col_idx - 1][2]

                wb.save(str(file_path))

                logger.info(
                    "Fiş aktarım satır(lar)ı eklendi",
                    event="excel_row_added",
                    file=file_path.name,
                    first_row=first_row,
                    row_count=len(template_rows),
                    store=data.firma,
                    total=data.toplam,
                    source=source,
                    confidence=confidence,
                )
                return first_row

            except Exception as e:
                logger.error("Excel yazma hatası", event="excel_error", error=str(e))
                raise

    async def update_row(self, row_number: int, data: dict, target_date: Optional[date] = None) -> bool:
        """Belirtilen satır grubunu güncelle (aynı Fiş No'ya sahip satırlar).

        Desteklenen alanlar: fis_no, tarih, firma, masraf, toplam, odeme, kdv_oran, kdv_tutar, matrah.
        KDV/matrah/toplam değiştiğinde ilgili muhasebe satırları tam olarak yeniden hesaplanır.
        """
        async with self._lock:
            try:
                file_path = self._daily_path(target_date)
                if not file_path.exists():
                    return False

                wb = load_workbook(str(file_path))
                if self.SHEET_NAME not in wb.sheetnames:
                    wb.close()
                    return False
                ws = wb[self.SHEET_NAME]

                if row_number < 2 or row_number > ws.max_row:
                    wb.close()
                    return False

                fis_no = ws.cell(row=row_number, column=1).value
                group_rows = [row_number]
                if fis_no:
                    for r in range(row_number + 1, ws.max_row + 1):
                        if ws.cell(row=r, column=1).value == fis_no:
                            group_rows.append(r)
                        else:
                            break

                # ── Tüm satırlarda ortak güncelleme: fis_no, tarih, firma ──
                for r in group_rows:
                    if "fis_no" in data and data["fis_no"] is not None:
                        ws.cell(row=r, column=1, value=data["fis_no"])
                        ws.cell(row=r, column=5, value=data["fis_no"])

                    if "tarih" in data and data["tarih"] is not None:
                        formatted = self._parse_receipt_date(data["tarih"])
                        ws.cell(row=r, column=2, value=formatted)
                        ws.cell(row=r, column=6, value=formatted)

                    if "firma" in data and data["firma"] is not None:
                        ws.cell(row=r, column=3, value=data["firma"])

                # ── Satır türlerini tanımla ──
                masraf_rows = []  # 770.xx hesap kodlu satırlar (gider)
                kdv_rows = []     # 191.xx hesap kodlu satırlar
                kkeg_borc_rows = []   # 900 hesap kodlu satırlar
                kkeg_alacak_rows = [] # 901 hesap kodlu satırlar
                odeme_rows = []   # 100.xx/102.xx hesap kodlu satırlar (son satır)

                for r in group_rows:
                    hesap = str(ws.cell(row=r, column=4).value or "")
                    if hesap.startswith("770"):
                        masraf_rows.append(r)
                    elif hesap.startswith("191"):
                        kdv_rows.append(r)
                    elif hesap == "900":
                        kkeg_borc_rows.append(r)
                    elif hesap == "901":
                        kkeg_alacak_rows.append(r)
                    elif hesap.startswith("100") or hesap.startswith("102"):
                        odeme_rows.append(r)

                has_kkeg = len(kkeg_borc_rows) > 0
                firma_val = data.get("firma") or ws.cell(row=group_rows[0], column=3).value or ""

                # ── Masraf türü güncellemesi ──
                if "masraf" in data and data["masraf"] is not None:
                    masraf_kodu = MASRAF_HESAP_KODU.get(data["masraf"], "770.99")
                    for r in masraf_rows:
                        ws.cell(row=r, column=4, value=masraf_kodu)
                        ws.cell(row=r, column=7, value=f"{data['masraf']} Gideri - {firma_val}")

                # ── Ödeme türü güncellemesi ──
                if "odeme" in data and data["odeme"] is not None:
                    odeme_key = data["odeme"].upper()
                    odeme_kodu = ODEME_HESAP_KODU.get(odeme_key, "100.01")
                    belge_tr = ODEME_BELGE_TR.get(odeme_key, "Fiş")
                    for r in odeme_rows:
                        ws.cell(row=r, column=4, value=odeme_kodu)
                    for r in group_rows:
                        ws.cell(row=r, column=11, value=belge_tr)

                # ── KDV/Matrah/Toplam güncellemesi ──
                needs_recalc = any(
                    k in data and data[k] is not None
                    for k in ("toplam", "matrah", "kdv_tutar", "kdv_oran")
                )

                if needs_recalc:
                    toplam = float(data["toplam"]) if ("toplam" in data and data["toplam"] is not None) else None
                    matrah = float(data["matrah"]) if ("matrah" in data and data["matrah"] is not None) else None
                    kdv_tutar = float(data["kdv_tutar"]) if ("kdv_tutar" in data and data["kdv_tutar"] is not None) else None
                    kdv_oran = data.get("kdv_oran")

                    if toplam is None:
                        if has_kkeg and odeme_rows and kkeg_alacak_rows:
                            # KKEG durumunda: ödeme Alacak = matrah + kdv_70
                            # Gerçek toplam = ödeme Alacak + KKEG Alacak toplamı
                            odeme_val = float(ws.cell(row=odeme_rows[-1], column=9).value or 0)
                            kkeg_alacak_val = sum(float(ws.cell(row=r, column=9).value or 0) for r in kkeg_alacak_rows)
                            toplam = odeme_val + kkeg_alacak_val
                        elif odeme_rows:
                            toplam = float(ws.cell(row=odeme_rows[-1], column=9).value or 0)
                    if matrah is None:
                        if has_kkeg and masraf_rows and kkeg_borc_rows:
                            # KKEG durumunda: masraf satırı Borç = matrah - kdv_30
                            # Gerçek matrah = masraf Borç + KKEG Borç toplamı
                            gider_val = float(ws.cell(row=masraf_rows[0], column=8).value or 0)
                            kkeg_val = sum(float(ws.cell(row=r, column=8).value or 0) for r in kkeg_borc_rows)
                            matrah = gider_val + kkeg_val
                        elif masraf_rows:
                            matrah = float(ws.cell(row=masraf_rows[0], column=8).value or 0)
                    if kdv_tutar is None and kdv_rows:
                        kdv_tutar = sum(float(ws.cell(row=r, column=8).value or 0) for r in kdv_rows)

                    if toplam is not None and matrah is None and kdv_tutar is None:
                        if kdv_oran is None and kdv_rows:
                            detay = str(ws.cell(row=kdv_rows[0], column=7).value or "")
                            m = re.search(r'%(\d+)', detay)
                            if m:
                                kdv_oran = f"%{m.group(1)}"

                        if kdv_oran:
                            m = re.search(r'(\d+)', str(kdv_oran))
                            if m:
                                oran_decimal = int(m.group(1)) / 100.0
                                matrah = round(toplam / (1 + oran_decimal), 2)
                                kdv_tutar = round(toplam - matrah, 2)

                    if kdv_oran and kdv_rows:
                        oran_key = kdv_oran if kdv_oran.startswith('%') else f"%{kdv_oran}"
                        m = re.search(r'(\d+)', str(oran_key))
                        if m:
                            oran_key = f"%{m.group(1)}"
                        kdv_kodu = KDV_HESAP_KODU.get(oran_key, f"191.{oran_key.replace('%', '').zfill(2)}")
                        for r in kdv_rows:
                            ws.cell(row=r, column=4, value=kdv_kodu)
                            old_detay = str(ws.cell(row=r, column=7).value or "")
                            new_detay = re.sub(r'%\d+', oran_key, old_detay) if '%' in old_detay else f"KDV {oran_key} - {firma_val}"
                            ws.cell(row=r, column=7, value=new_detay)

                    # ── KKEG (70/30) durumunda yeniden hesapla ──
                    if has_kkeg and matrah is not None and kdv_tutar is not None:
                        kdv_70 = round(kdv_tutar * 0.70, 2)
                        kdv_30 = round(kdv_tutar - kdv_70, 2)
                        gider_tutar = round(matrah - kdv_30, 2)
                        odeme_tutar = round(matrah + kdv_70, 2)

                        # Masraf satırı (gider)
                        for r in masraf_rows:
                            ws.cell(row=r, column=8, value=gider_tutar)
                        # KDV satırları (%70)
                        for r in kdv_rows:
                            ws.cell(row=r, column=8, value=kdv_70)
                        # KKEG Borç (900)
                        for r in kkeg_borc_rows:
                            ws.cell(row=r, column=8, value=kdv_30)
                        # KKEG Alacak (901)
                        for r in kkeg_alacak_rows:
                            ws.cell(row=r, column=9, value=kdv_30)
                        # Ödeme satırı
                        for r in odeme_rows:
                            ws.cell(row=r, column=9, value=odeme_tutar)

                    # ── Normal fiş durumunda yeniden hesapla ──
                    elif not has_kkeg:
                        if matrah is not None and masraf_rows:
                            for r in masraf_rows:
                                ws.cell(row=r, column=8, value=round(matrah, 2))

                        if kdv_tutar is not None and kdv_rows:
                            # Tek KDV satırı varsa tamamını yaz
                            if len(kdv_rows) == 1:
                                ws.cell(row=kdv_rows[0], column=8, value=round(kdv_tutar, 2))

                        if toplam is not None and odeme_rows:
                            for r in odeme_rows:
                                ws.cell(row=r, column=9, value=round(toplam, 2))

                wb.save(str(file_path))
                wb.close()

                logger.info(
                    "Excel satır grubu güncellendi",
                    event="excel_row_updated",
                    file=file_path.name,
                    row_number=row_number,
                    group_size=len(group_rows),
                    fields=list(data.keys()),
                )
                return True

            except Exception as e:
                logger.error("Excel güncelleme hatası", event="excel_update_error", error=str(e))
                raise

    async def get_row_count(self, target_date: Optional[date] = None) -> int:
        async with self._lock:
            try:
                path = self._daily_path(target_date)
                if not path.exists():
                    return 0
                wb = load_workbook(str(path), read_only=True)
                sheet = self.SHEET_NAME if self.SHEET_NAME in wb.sheetnames else wb.sheetnames[0]
                ws = wb[sheet]
                count = max(ws.max_row - 1, 0)
                wb.close()
                return count
            except Exception:
                return 0

    async def delete_row(self, row_number: int, target_date: Optional[date] = None) -> dict:
        """Belirtilen satır grubunu (aynı Fiş No) Excel'den siler."""
        async with self._lock:
            try:
                file_path = self._daily_path(target_date)
                if not file_path.exists():
                    return {"deleted": 0, "error": "Dosya bulunamadı"}

                wb = load_workbook(str(file_path))
                if self.SHEET_NAME not in wb.sheetnames:
                    wb.close()
                    return {"deleted": 0, "error": "Sayfa bulunamadı"}
                ws = wb[self.SHEET_NAME]

                if row_number < 2 or row_number > ws.max_row:
                    wb.close()
                    return {"deleted": 0, "error": "Geçersiz satır numarası"}

                fis_no = ws.cell(row=row_number, column=1).value
                group_rows = [row_number]
                if fis_no:
                    for r in range(row_number + 1, ws.max_row + 1):
                        if ws.cell(row=r, column=1).value == fis_no:
                            group_rows.append(r)
                        else:
                            break

                deleted_count = len(group_rows)
                for r in sorted(group_rows, reverse=True):
                    ws.delete_rows(r, 1)

                wb.save(str(file_path))
                wb.close()

                logger.info(
                    "Excel satır grubu silindi",
                    event="excel_row_deleted",
                    file=file_path.name,
                    row_number=row_number,
                    group_size=deleted_count,
                )
                return {"deleted": deleted_count, "first_row": row_number}

            except Exception as e:
                logger.error("Excel silme hatası", event="excel_delete_error", error=str(e))
                raise

    async def get_file_path(self, target_date: Optional[date] = None) -> Optional[str]:
        path = self._daily_path(target_date)
        if path.exists():
            return str(path.resolve())
        return None

    async def get_file_bytes(self, target_date: Optional[date] = None) -> Optional[bytes]:
        async with self._lock:
            path = self._daily_path(target_date)
            if path.exists():
                return path.read_bytes()
            return None

    async def export_as_xlsx(self, target_date: Optional[date] = None) -> Optional[str]:
        """XLSX dosya yolu döndür (zaten şablon formatında)."""
        return await self.get_file_path(target_date)

    async def export_as_csv(self, target_date: Optional[date] = None) -> Optional[bytes]:
        """Şablon formatında CSV çıktısı üret (UTF-8 BOM, ; ayraçlı)."""
        async with self._lock:
            path = self._daily_path(target_date)
            if not path.exists():
                return None
            try:
                wb = load_workbook(str(path), read_only=True)
                sheet = self.SHEET_NAME if self.SHEET_NAME in wb.sheetnames else wb.sheetnames[0]
                ws = wb[sheet]

                output = io.StringIO()
                writer = csv.writer(output, delimiter=";", quoting=csv.QUOTE_MINIMAL)
                for row in ws.iter_rows(values_only=True):
                    writer.writerow([v if v is not None else "" for v in row])

                wb.close()
                return ("\ufeff" + output.getvalue()).encode("utf-8")
            except Exception as e:
                logger.error("CSV export hatası", event="csv_export_error", error=str(e))
                return None

    async def export_as_xls(self, target_date: Optional[date] = None) -> Optional[bytes]:
        """Legacy .xls (BIFF) formatında çıktı üret."""
        async with self._lock:
            path = self._daily_path(target_date)
            if not path.exists():
                return None
            try:
                import xlwt

                wb_src = load_workbook(str(path), read_only=True)
                sheet = self.SHEET_NAME if self.SHEET_NAME in wb_src.sheetnames else wb_src.sheetnames[0]
                ws_src = wb_src[sheet]

                wb_xls = xlwt.Workbook(encoding="utf-8")
                ws_xls = wb_xls.add_sheet(self.SHEET_NAME)

                header_style = xlwt.easyxf(
                    "font: name Calibri, bold on, height 200;"
                    "pattern: pattern solid, fore_colour gray25;"
                    "align: horiz center, vert center, wrap on;"
                )
                money_fmt = xlwt.easyxf(num_format_str="#,##0.00")

                for row_idx, row in enumerate(ws_src.iter_rows(values_only=True)):
                    for col_idx, value in enumerate(row):
                        val = value if value is not None else ""
                        if row_idx == 0:
                            ws_xls.write(row_idx, col_idx, val, header_style)
                        elif col_idx in (7, 8, 13):
                            ws_xls.write(row_idx, col_idx, val, money_fmt)
                        else:
                            ws_xls.write(row_idx, col_idx, val)

                for col_idx, (_, width, _) in enumerate(self.TEMPLATE_COLUMNS):
                    ws_xls.col(col_idx).width = int(width * 256)

                wb_src.close()
                buf = io.BytesIO()
                wb_xls.save(buf)
                return buf.getvalue()

            except ImportError:
                logger.warning("xlwt yüklü değil, XLS export devre dışı")
                return None
            except Exception as e:
                logger.error("XLS export hatası", event="xls_export_error", error=str(e))
                return None

    async def export_all_combined(self, fmt: str = "xlsx") -> Optional[str | bytes]:
        """Tüm günlük dosyaları birleştirir. fmt: xlsx | csv | xls"""
        async with self._lock:
            try:
                daily_files = sorted(self._data_dir.glob("*.xlsx"))
                if not daily_files:
                    return None

                combined_wb = Workbook()
                combined_ws = combined_wb.active
                combined_ws.title = self.SHEET_NAME
                self._setup_header(combined_ws)

                combined_row = 2
                total_files = 0

                for daily_file in daily_files:
                    try:
                        wb = load_workbook(str(daily_file), read_only=True)
                        sheet = self.SHEET_NAME if self.SHEET_NAME in wb.sheetnames else wb.sheetnames[0]
                        ws = wb[sheet]
                        for row in ws.iter_rows(min_row=2, values_only=True):
                            if not any(row):
                                continue
                            for col_idx, value in enumerate(row, 1):
                                cell = combined_ws.cell(row=combined_row, column=col_idx, value=value)
                                if col_idx <= len(self.TEMPLATE_COLUMNS):
                                    cell.number_format = self.TEMPLATE_COLUMNS[col_idx - 1][2]
                            combined_row += 1
                        wb.close()
                        total_files += 1
                    except Exception as e:
                        logger.warning(f"Dosya okunamadı: {daily_file.name}", error=str(e))

                self._create_combined_summary(combined_wb, combined_row - 2, total_files)

                output_path = self._data_dir.parent / "all_invoices.xlsx"
                combined_wb.save(str(output_path))

                logger.info(
                    "Birleşik dosya oluşturuldu",
                    event="excel_combined",
                    total_files=total_files,
                    total_rows=combined_row - 2,
                    format=fmt,
                )

                if fmt == "csv":
                    return self._xlsx_to_csv_bytes(output_path)
                elif fmt == "xls":
                    return self._xlsx_to_xls_bytes(output_path)
                else:
                    return str(output_path.resolve())

            except Exception as e:
                logger.error("Birleştirme hatası", event="excel_combine_error", error=str(e))
                raise

    def _xlsx_to_csv_bytes(self, xlsx_path: Path) -> Optional[bytes]:
        try:
            wb = load_workbook(str(xlsx_path), read_only=True)
            sheet = self.SHEET_NAME if self.SHEET_NAME in wb.sheetnames else wb.sheetnames[0]
            ws = wb[sheet]
            output = io.StringIO()
            writer = csv.writer(output, delimiter=";", quoting=csv.QUOTE_MINIMAL)
            for row in ws.iter_rows(values_only=True):
                writer.writerow([v if v is not None else "" for v in row])
            wb.close()
            return ("\ufeff" + output.getvalue()).encode("utf-8")
        except Exception as e:
            logger.error("XLSX→CSV dönüşüm hatası", error=str(e))
            return None

    def _xlsx_to_xls_bytes(self, xlsx_path: Path) -> Optional[bytes]:
        try:
            import xlwt
            wb_src = load_workbook(str(xlsx_path), read_only=True)
            sheet = self.SHEET_NAME if self.SHEET_NAME in wb_src.sheetnames else wb_src.sheetnames[0]
            ws_src = wb_src[sheet]

            wb_xls = xlwt.Workbook(encoding="utf-8")
            ws_xls = wb_xls.add_sheet(self.SHEET_NAME)

            header_style = xlwt.easyxf(
                "font: name Calibri, bold on, height 200;"
                "pattern: pattern solid, fore_colour gray25;"
                "align: horiz center, vert center, wrap on;"
            )
            for row_idx, row in enumerate(ws_src.iter_rows(values_only=True)):
                for col_idx, val in enumerate(row):
                    v = val if val is not None else ""
                    ws_xls.write(row_idx, col_idx, v, header_style if row_idx == 0 else xlwt.Style.default_style)

            for col_idx, (_, width, _) in enumerate(self.TEMPLATE_COLUMNS):
                ws_xls.col(col_idx).width = int(width * 256)

            wb_src.close()
            buf = io.BytesIO()
            wb_xls.save(buf)
            return buf.getvalue()
        except ImportError:
            logger.warning("xlwt yüklü değil, XLS export devre dışı")
            return None
        except Exception as e:
            logger.error("XLSX→XLS dönüşüm hatası", error=str(e))
            return None

    def _create_combined_summary(self, wb: Workbook, total_rows: int, total_files: int):
        ws = wb.create_sheet("Özet")
        title_font = Font(name="Calibri", bold=True, size=14, color="1F4E79")
        label_font = Font(name="Calibri", bold=True, size=11)
        value_font = Font(name="Calibri", size=11)

        ws.column_dimensions["A"].width = 28
        ws.column_dimensions["B"].width = 30

        ws["A1"] = "Fatura Bot — Fiş Aktarım Özeti"
        ws["A1"].font = title_font

        sheet_ref = f"'{self.SHEET_NAME}'"
        summary = [
            ("A3", "Toplam Gün Sayısı:", "B3", total_files),
            ("A4", "Toplam Satır Sayısı:", "B4", f"=COUNTA({sheet_ref}!A2:A1000000)"),
            ("A5", "Toplam Borç (₺):", "B5", f"=SUM({sheet_ref}!H2:H1000000)"),
            ("A6", "Toplam Alacak (₺):", "B6", f"=SUM({sheet_ref}!I2:I1000000)"),
            ("A7", "Oluşturma Tarihi:", "B7", datetime.now(timezone.utc).strftime("%d/%m/%Y %H:%M")),
        ]
        for label_ref, label, val_ref, value in summary:
            ws[label_ref] = label
            ws[label_ref].font = label_font
            ws[val_ref] = value
            ws[val_ref].font = value_font
        ws["B5"].number_format = "#,##0.00"
        ws["B6"].number_format = "#,##0.00"

    async def list_daily_files(self) -> list[dict]:
        files = []
        for f in sorted(self._data_dir.glob("*.xlsx"), reverse=True):
            try:
                wb = load_workbook(str(f), read_only=True)
                sheet = self.SHEET_NAME if self.SHEET_NAME in wb.sheetnames else wb.sheetnames[0]
                ws = wb[sheet]
                count = max(ws.max_row - 1, 0)
                wb.close()
                files.append({"date": f.stem, "file": f.name, "row_count": count})
            except Exception:
                files.append({"date": f.stem, "file": f.name, "row_count": 0})
        return files

    async def read_queries_from_excel(self, target_date: Optional[date] = None, limit: int = 50) -> list[dict]:
        """Excel dosyasından sorgu satırlarını okur — bellekten bağımsız, restart'a dayanıklı."""
        async with self._lock:
            try:
                path = self._daily_path(target_date)
                if not path.exists():
                    return []

                wb = load_workbook(str(path), read_only=True)
                sheet = self.SHEET_NAME if self.SHEET_NAME in wb.sheetnames else wb.sheetnames[0]
                ws = wb[sheet]

                # Tüm satırları oku, aynı fiş no'ya sahip satırları grupla
                rows_data = []
                for row in ws.iter_rows(min_row=2, values_only=False):
                    vals = [cell.value for cell in row]
                    if not any(vals):
                        continue
                    rows_data.append(vals)

                wb.close()

                if not rows_data:
                    return []

                # Fiş No bazında grupla — her fiş grubunun ilk satırı ana veriyi taşır
                queries = []
                i = 0
                file_date = (target_date or date.today()).isoformat()
                while i < len(rows_data):
                    row = rows_data[i]
                    fis_no = row[0] if len(row) > 0 else None
                    tarih = row[1] if len(row) > 1 else None
                    firma = row[2] if len(row) > 2 else None
                    hesap_kodu = str(row[3]) if len(row) > 3 and row[3] else ""
                    detay = str(row[6]) if len(row) > 6 and row[6] else ""
                    borc = float(row[7]) if len(row) > 7 and row[7] else 0
                    alacak = float(row[8]) if len(row) > 8 and row[8] else 0

                    # Grup satırlarını topla (aynı fiş no)
                    group_start = i
                    group_borc_total = borc
                    group_alacak_total = alacak
                    kdv_tutar = 0
                    matrah = 0
                    kdv_oran = ""
                    masraf = ""
                    odeme = ""
                    toplam = 0

                    # Ana satır masraf türü tespit
                    if hesap_kodu.startswith("770"):
                        masraf = detay.split(" Gideri")[0].strip() if " Gideri" in detay else detay
                        matrah = borc

                    j = i + 1
                    while j < len(rows_data) and rows_data[j][0] == fis_no:
                        r = rows_data[j]
                        r_hesap = str(r[3]) if len(r) > 3 and r[3] else ""
                        r_detay = str(r[6]) if len(r) > 6 and r[6] else ""
                        r_borc = float(r[7]) if len(r) > 7 and r[7] else 0
                        r_alacak = float(r[8]) if len(r) > 8 and r[8] else 0
                        group_borc_total += r_borc
                        group_alacak_total += r_alacak

                        if r_hesap.startswith("191"):
                            kdv_tutar += r_borc
                            # KDV oranını detaydan çıkar
                            import re
                            m = re.search(r'%\d+', r_detay)
                            if m:
                                kdv_oran = m.group(0)
                        elif r_hesap.startswith("100") or r_hesap.startswith("102"):
                            toplam = r_alacak
                            if r_hesap.startswith("100"):
                                odeme = "NAKİT"
                            else:
                                odeme = "KART"
                        elif r_hesap.startswith("770") and not masraf:
                            masraf = r_detay.split(" Gideri")[0].strip() if " Gideri" in r_detay else r_detay
                            matrah = r_borc

                        j += 1

                    if toplam == 0:
                        toplam = group_alacak_total

                    # Belge türünden ödeme tespiti
                    belge_tr = str(rows_data[i][10]) if len(rows_data[i]) > 10 and rows_data[i][10] else ""
                    if not odeme:
                        if "Kasa" in belge_tr or "Fiş" in belge_tr:
                            odeme = "NAKİT"
                        elif "Banka" in belge_tr or "Dekont" in belge_tr:
                            odeme = "KART"

                    queries.append({
                        "request_id": f"excel_{file_date}_{group_start + 2}",
                        "timestamp": f"{file_date}T00:00:00Z",
                        "firma": firma,
                        "toplam": round(toplam, 2),
                        "confidence": 100,
                        "source": "excel",
                        "processing_time_ms": 0,
                        "masraf": masraf,
                        "tarih": tarih,
                        "odeme": odeme,
                        "fis_no": fis_no,
                        "vkn": None,
                        "matrah": round(matrah, 2),
                        "kdv_oran": kdv_oran,
                        "kdv_tutar": round(kdv_tutar, 2),
                        "status": "success",
                        "row_number": group_start + 2,
                        "file_date": file_date,
                    })

                    i = j

                queries.reverse()  # En yeni en üstte
                return queries[:limit]

            except Exception as e:
                logger.error("Excel sorgu okuma hatası", event="excel_read_queries_error", error=str(e))
                return []

